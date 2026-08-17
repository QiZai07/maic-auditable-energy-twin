from __future__ import annotations

import hashlib
import json
import math
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import linprog
from scipy.sparse import lil_matrix
from sklearn.ensemble import HistGradientBoostingRegressor
from sklearn.metrics import mean_absolute_error, r2_score


plt.rcParams["font.sans-serif"] = ["Arial", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False


BUILDING = {
    "building_id": "BLD_REF",
    "official_name": "Ningbo Reference Building",
    "short_name": "NRB",
    "source_alias": "Reference Building A",
    "institution": "Anonymized Ningbo Campus",
    "city": "Ningbo",
    "country": "China",
    "target_market": "Malaysia",
    "gross_floor_area_m2": 6231.26,
    "floor_count": 3,
    "floor_areas_m2": {"1F": 2047.88, "2F": 2091.69, "3F": 2091.69},
    "identity_status": "anonymized_public_reference_case",
}

TARIFF = {
    "value": 0.538,
    "currency": "CNY",
    "unit": "CNY/kWh",
    "tax_included": True,
    "billing_structure": "single_flat_energy_rate",
    "billing_formula": "bill_cny = electricity_kwh * 0.538",
    "time_of_use_charge": False,
    "demand_charge": False,
    "other_charge_components": False,
    "status": "Ningbo_reference_case_parameter",
}

MAIC_EMISSION_FACTOR = {
    "value_kgco2e_per_kwh": 0.740,
    "boundary": "Malaysia deployment scenario only",
    "status": "competition_scenario_assumption",
}

MODEL_VERSION = "Irene Auditable Digital Twin"
OPENING_HOUR = 8
CLOSING_HOUR = 22

COMFORT_ASSUMPTION = {
    "status": "user_confirmed_model_assumption_2026-08-10",
    "scope": "classrooms during 08:00-22:00 opening hours",
    "cooling_target_c": 25.0,
    "heating_target_c": 20.0,
    "acceptable_temperature_band_c": [20.0, 26.0],
    "relative_humidity_band_pct": [40.0, 60.0],
    "co2_upper_limit_ppm": 1000.0,
    "evidence_boundary": "synthetic comfort proxy; no historical indoor sensor records",
}

HVAC_DESIGN_BASIS = {
    "source": "Anonymized reference design basis; source drawing excluded",
    "outdoor_unit_count": 18,
    "indoor_unit_count": 80,
    "rated_cooling_capacity_kw": 952.0,
    "rated_heating_capacity_kw": 1066.0,
    "rated_input_power_kw": 265.0,
    "rated_cooling_eer": round(952.0 / 265.0, 4),
    "rated_heating_cop": round(1066.0 / 265.0, 4),
    "fresh_air_fan_power_kw": 6.0,
    "general_exhaust_fan_power_kw": 4.1,
    "status": "design_basis_not_current_nameplate_measurement",
}

CURRENT_PV = {
    "installed": True,
    "grid_connection_point": "Grid connection point 2",
    "capacity_kwp": 106.14,
    "inverter_capacity_kw": 110.0,
    "module_count": 183,
    "module_power_wp": 580.0,
    "storage_installed": False,
    "status": "approved_aggregate_reference_case",
}


class CalendarShapePredictor:
    """Validated fallback that predicts normalized load shape by weekend/hour."""

    def __init__(self, lookup: dict[tuple[bool, int], float], default_value: float):
        self.lookup = lookup
        self.default_value = float(default_value)

    def predict(self, features: pd.DataFrame) -> np.ndarray:
        return np.array(
            [
                self.lookup.get((bool(is_weekend), int(hour)), self.default_value)
                for is_weekend, hour in zip(features["is_weekend"], features["hour_numeric"], strict=True)
            ],
            dtype=float,
        )


@dataclass(frozen=True)
class ProjectPaths:
    project_root: Path
    raw_meter_workbook: Path
    external_snapshots: Path
    processed_dir: Path
    config_dir: Path
    model_dir: Path
    reports_dir: Path
    charts_dir: Path
    app_dir: Path
    results_dir: Path
    logs_dir: Path
    docs_dir: Path
    test_dir: Path
    submission_dir: Path
    db_pv_workbook: Path


def get_paths(project_root: Path) -> ProjectPaths:
    project_root = project_root.resolve()
    return ProjectPaths(
        project_root=project_root,
        raw_meter_workbook=project_root / "data" / "private_inputs" / "Private_Monthly_Meter_Readings.xlsx",
        external_snapshots=project_root / "data" / "private_inputs",
        processed_dir=project_root / "data" / "processed",
        config_dir=project_root / "data" / "config",
        model_dir=project_root / "data" / "models",
        reports_dir=project_root / "reports",
        charts_dir=project_root / "charts",
        app_dir=project_root,
        results_dir=project_root / "results",
        logs_dir=project_root / "logs",
        docs_dir=project_root / "docs",
        test_dir=project_root / "tests",
        submission_dir=project_root / "submission",
        db_pv_workbook=project_root / "data" / "private_inputs" / "DB_PV_Monthly.xlsx",
    )


def ensure_directories(paths: ProjectPaths) -> None:
    for directory in (
        paths.external_snapshots,
        paths.processed_dir,
        paths.config_dir,
        paths.model_dir,
        paths.reports_dir,
        paths.charts_dir,
        paths.results_dir,
        paths.logs_dir,
        paths.test_dir,
    ):
        directory.mkdir(parents=True, exist_ok=True)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_json(data: Any, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def write_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")


def copy_external_snapshots(source_root: Path, paths: ProjectPaths, refresh: bool = False) -> pd.DataFrame:
    source_root = source_root.resolve()
    candidates = {
        "Ningbo_reference.epw": source_root / "Ningbo_reference.epw",
        "CEST.csv": source_root / "CEST.csv",
        "Private_Weather_Hourly.csv": source_root / "Private_Weather_Hourly.csv",
        "Weather_2023_5min.csv": source_root / "Weather_2023_5min.csv",
        "Private_Reference_Load.xlsx": source_root / "Private_Reference_Load.xlsx",
        "Private_Reference_PV.xlsx": source_root / "Private_Reference_PV.xlsx",
        "DB_PV_Monthly.xlsx": source_root / "DB_PV_Monthly.xlsx",
        "DB_Data_Responses.xlsx": source_root / "DB_Data_Responses.xlsx",
        "DB_PV_AsBuilt.dwg": source_root / "DB_PV_AsBuilt.dwg",
    }
    records: list[dict[str, Any]] = []
    for snapshot_name, source in candidates.items():
        if not source.exists():
            records.append(
                {
                    "snapshot_name": snapshot_name,
                    "source_path": str(source),
                    "status": "missing",
                    "sha256": "",
                    "size_bytes": 0,
                    "evidence_type": "missing",
                }
            )
            continue
        target = paths.external_snapshots / snapshot_name
        if refresh or not target.exists() or sha256_file(source) != sha256_file(target):
            shutil.copy2(source, target)
            status = "copied"
        else:
            status = "unchanged"
        records.append(
            {
                "snapshot_name": snapshot_name,
                "source_path": str(source),
                "snapshot_path": str(target),
                "status": status,
                "sha256": sha256_file(target),
                "size_bytes": target.stat().st_size,
                "evidence_type": "read_only_source_snapshot",
            }
        )
    return pd.DataFrame(records)


def source_lineage(paths: ProjectPaths) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    source_specs = [
        (paths.raw_meter_workbook, "reference case monthly meter readings", "measured", "Monthly building and submeter energy anchor"),
        (paths.external_snapshots / "Private_Building_Profile.docx", "authorized building profile", "documented", "Name, floors and area"),
        (paths.external_snapshots / "Private_Weather_Hourly.csv", "reference weather 2024", "measured", "Hourly weather for 2024"),
        (paths.external_snapshots / "Weather_2023_5min.csv", "reference weather 2023", "measured", "Climatology support for missing 2025 weather"),
        (paths.external_snapshots / "Private_Reference_Load.xlsx", "reference donor energy building", "measured_reference", "Shape transfer only; absolute kWh is not reference case data"),
        (paths.external_snapshots / "Private_Reference_PV.xlsx", "reference reference PV", "measured_reference", "Intraday shape only for reference case monthly-constrained PV estimation"),
        (paths.db_pv_workbook, "reference case PV monthly generation", "measured", "Monthly generation anchor for the installed 106.14 kWp system"),
        (paths.external_snapshots / "DB_Data_Responses.xlsx", "Facilities data responses", "documented", "Confirms 25 C setpoints, no storage, meter fault and missing BMS/indoor records"),
        (paths.external_snapshots / "Private_PV_AsBuilt.dwg", "authorized PV as-built drawing", "as_built_document", "Confirms installed PV equipment"),
        (paths.external_snapshots / "Private_HVAC_Design.dwg", "authorized HVAC design drawing", "design_document", "Equipment schedule and rated performance; not current measured performance"),
        (paths.external_snapshots / "Private_Electrical_Design.dwg", "authorized electrical design drawing", "design_document", "Panel and circuit design basis"),
        (paths.external_snapshots / "Private_Architectural_Design.dwg", "authorized architectural design drawing", "design_document", "Envelope and zoning prior"),
    ]
    for path, source_name, evidence_type, use_boundary in source_specs:
        rows.append(
            {
                "source_name": source_name,
                "file_name": path.name,
                "path": str(path),
                "exists": path.exists(),
                "sha256": sha256_file(path) if path.exists() else "",
                "size_bytes": path.stat().st_size if path.exists() else 0,
                "evidence_type": evidence_type if path.exists() else "missing",
                "use_boundary": use_boundary,
            }
        )
    return pd.DataFrame(rows)


def _to_float(value: Any) -> float:
    if value is None:
        return float("nan")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.startswith("="):
        return float("nan")
    return float(text)


def build_monthly_meter_dataset(workbook_path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    data_rows = [list(row) for row in rows[1:] if any(cell is not None for cell in row)]
    months = [
        ("2024-07", 4, 5, True),
        ("2024-08", 6, 7, False),
        ("2024-09", 8, 9, False),
        ("2024-10", 10, 11, False),
        ("2024-11", 12, 13, False),
        ("2024-12", 14, 15, False),
        ("2025-01", 16, 17, False),
        ("2025-02", 18, 19, False),
        ("2025-03", 20, 21, False),
        ("2025-04", 22, 23, False),
        ("2025-05", 24, 25, False),
        ("2025-06", 26, 27, False),
    ]
    monthly_records: list[dict[str, Any]] = []
    summary_records: list[dict[str, Any]] = []
    source_label = ""
    source_tariff = float("nan")
    for source_row, row in enumerate(data_rows, start=2):
        source_label = source_label or str(row[0]).strip()
        meter_name_raw = str(row[2]).strip()
        meter_code = next((code for code in ("MTR-A", "MTR-B", "MTR-C", "MTR-D") if code in meter_name_raw.upper()), meter_name_raw)
        meter_name = f"{meter_code} meter (rooftop HVAC)" if meter_code in {"MTR-C", "MTR-D"} else f"{meter_code} meter"
        ratio = _to_float(row[3])
        source_tariff = source_tariff if math.isfinite(source_tariff) else _to_float(row[29])
        meter_group = "Rooftop HVAC" if meter_code in {"MTR-C", "MTR-D"} else "Non-HVAC"
        usage_map: dict[str, float] = {}
        zero_increment_months: list[str] = []
        for month, cumulative_index, usage_index, direct_usage in months:
            current = _to_float(row[cumulative_index])
            if direct_usage:
                usage = _to_float(row[usage_index])
            else:
                previous = _to_float(row[cumulative_index - 2])
                usage = (current - previous) * ratio
                if math.isclose(current, previous, abs_tol=1e-12):
                    zero_increment_months.append(month)
            usage_map[month] = max(float(usage), 0.0)
            monthly_records.append(
                {
                    "building_id": BUILDING["building_id"],
                    "building_name": BUILDING["official_name"],
                    "building_short_name": BUILDING["short_name"],
                    "source_building_label": source_label,
                    "meter_name": meter_name,
                    "meter_group": meter_group,
                    "month": month,
                    "usage_kwh": round(max(float(usage), 0.0), 6),
                    "tariff_cny_per_kwh": TARIFF["value"],
                    "source_row": source_row,
                    "evidence_type": "measured_or_meter_difference",
                }
            )
        summary_records.append(
            {
                "building_id": BUILDING["building_id"],
                "meter_name": meter_name,
                "meter_group": meter_group,
                "multiplier": ratio,
                "annual_kwh": round(sum(usage_map.values()), 6),
                "zero_increment_months": ";".join(zero_increment_months),
                "source_row": source_row,
            }
        )
    monthly = pd.DataFrame(monthly_records)
    meter_summary = pd.DataFrame(summary_records)
    monthly_totals = monthly.groupby("month", as_index=False)["usage_kwh"].sum()
    monthly_totals["estimated_cost_cny"] = monthly_totals["usage_kwh"] * TARIFF["value"]
    monthly_totals["eui_kwh_m2"] = monthly_totals["usage_kwh"] / BUILDING["gross_floor_area_m2"]
    metadata = {
        "source_building_label": source_label,
        "canonical_building_id": BUILDING["building_id"],
        "canonical_building_name": BUILDING["official_name"],
        "source_tariff_value": source_tariff,
        "confirmed_tariff": TARIFF,
        "period": "2024-07 to 2025-06",
        "annual_total_kwh": float(monthly["usage_kwh"].sum()),
        "annual_hvac_kwh": float(monthly.loc[monthly["meter_group"] == "Rooftop HVAC", "usage_kwh"].sum()),
        "annual_non_hvac_kwh": float(monthly.loc[monthly["meter_group"] == "Non-HVAC", "usage_kwh"].sum()),
    }
    return monthly, meter_summary, monthly_totals, metadata


def load_weather_2024(path: Path) -> pd.DataFrame:
    raw = pd.read_csv(path)
    rename = {
        "Date Time": "timestamp",
        "Dry Bulb Temperature (°C)": "dry_bulb_c",
        "Wet Bulb Temperature (°C)": "wet_bulb_c",
        "Atmospheric Pressure (KPa)": "pressure_kpa",
        "Relative Humidity (%)": "relative_humidity_pct",
        "Dew Point Temperature (°C)": "dew_point_c",
        "GHI": "ghi_wh_m2",
        "DNI": "dni_wh_m2",
        "DHI": "dhi_wh_m2",
        "Wind Speed (m/s)": "wind_speed_m_s",
    }
    weather = raw.rename(columns=rename)[list(rename.values())].copy()
    weather["timestamp"] = pd.to_datetime(weather["timestamp"], errors="coerce")
    weather = weather.dropna(subset=["timestamp"]).drop_duplicates("timestamp").sort_values("timestamp")
    for column in weather.columns.drop("timestamp"):
        weather[column] = pd.to_numeric(weather[column], errors="coerce")
    weather["weather_evidence"] = "authorized_site_weather"
    return weather.reset_index(drop=True)


def load_weather_2023_hourly(path: Path) -> pd.DataFrame:
    raw = pd.read_csv(path, engine="python")
    weather = pd.DataFrame(
        {
            "timestamp": pd.to_datetime(raw["timestamp"], errors="coerce"),
            "dry_bulb_c": pd.to_numeric(raw["dry_bulb_c"], errors="coerce"),
            "relative_humidity_pct": pd.to_numeric(
                raw["relative_humidity_pct"], errors="coerce"
            ),
            "ghi_wh_m2": pd.to_numeric(
                raw["solar_irradiance_w_m2"], errors="coerce"
            ),
            "wind_speed_m_s": pd.to_numeric(
                raw["wind_speed_mph"], errors="coerce"
            )
            * 0.44704,
        }
    ).dropna(subset=["timestamp"])
    weather = weather.set_index("timestamp").resample("1h").mean(numeric_only=True).reset_index()
    weather["weather_evidence"] = "authorized_reference_weather_aggregated"
    return weather


def build_target_weather(weather_2024: pd.DataFrame, weather_2023: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    target_index = pd.date_range("2024-07-01 00:00:00", "2025-06-30 23:00:00", freq="1h")
    numeric = ["dry_bulb_c", "relative_humidity_pct", "ghi_wh_m2", "wind_speed_m_s"]
    w24 = weather_2024.set_index("timestamp")
    climatology_source = pd.concat(
        [
            weather_2023.assign(year_source=2023),
            weather_2024[numeric + ["timestamp"]].assign(year_source=2024),
        ],
        ignore_index=True,
    )
    climatology_source["month"] = climatology_source["timestamp"].dt.month
    climatology_source["day"] = climatology_source["timestamp"].dt.day
    climatology_source["hour"] = climatology_source["timestamp"].dt.hour
    climatology = climatology_source.groupby(["month", "day", "hour"], as_index=False)[numeric].mean()
    climate_lookup = climatology.set_index(["month", "day", "hour"])
    rows: list[dict[str, Any]] = []
    for timestamp in target_index:
        if timestamp.year == 2024 and timestamp in w24.index:
            record = {column: w24.at[timestamp, column] if column in w24.columns else np.nan for column in numeric}
            evidence = "authorized_site_weather"
        else:
            key = (timestamp.month, timestamp.day, timestamp.hour)
            climate = climate_lookup.loc[key] if key in climate_lookup.index else pd.Series(dtype=float)
            record = {column: climate.get(column, np.nan) for column in numeric}
            evidence = "climatology_2023_2024_for_missing_2025"
        rows.append({"timestamp": timestamp, **record, "weather_evidence": evidence})
    target = pd.DataFrame(rows)
    target["dry_bulb_c"] = target["dry_bulb_c"].interpolate(limit_direction="both")
    target["relative_humidity_pct"] = target["relative_humidity_pct"].interpolate(limit_direction="both")
    target["ghi_wh_m2"] = target["ghi_wh_m2"].fillna(0).clip(lower=0)
    target["wind_speed_m_s"] = target["wind_speed_m_s"].interpolate(limit_direction="both").clip(lower=0)
    return target, climatology


def load_donor_energy(path: Path) -> pd.DataFrame:
    donor = pd.read_excel(path, sheet_name=0)
    donor = donor.rename(columns={donor.columns[0]: "timestamp", "total_load_kwh": "total_load_kwh"})
    donor["timestamp"] = pd.to_datetime(donor["timestamp"], errors="coerce")
    donor["total_load_kwh"] = pd.to_numeric(donor["total_load_kwh"], errors="coerce")
    donor = donor.dropna(subset=["timestamp", "total_load_kwh"]).drop_duplicates("timestamp").sort_values("timestamp")
    donor.loc[donor["total_load_kwh"] < 0, "total_load_kwh"] = np.nan
    donor["total_load_kwh"] = donor["total_load_kwh"].interpolate(limit=3, limit_direction="both")
    return donor[["timestamp", "total_load_kwh"]].reset_index(drop=True)


def _calendar_weather_features(df: pd.DataFrame) -> pd.DataFrame:
    ts = pd.to_datetime(df["timestamp"])
    result = pd.DataFrame(index=df.index)
    result["hour_sin"] = np.sin(2 * np.pi * ts.dt.hour / 24)
    result["hour_cos"] = np.cos(2 * np.pi * ts.dt.hour / 24)
    result["hour_numeric"] = ts.dt.hour
    result["dow_sin"] = np.sin(2 * np.pi * ts.dt.dayofweek / 7)
    result["dow_cos"] = np.cos(2 * np.pi * ts.dt.dayofweek / 7)
    result["month_sin"] = np.sin(2 * np.pi * ts.dt.month / 12)
    result["month_cos"] = np.cos(2 * np.pi * ts.dt.month / 12)
    result["dayofyear_sin"] = np.sin(2 * np.pi * ts.dt.dayofyear / 365.25)
    result["dayofyear_cos"] = np.cos(2 * np.pi * ts.dt.dayofyear / 365.25)
    result["is_weekend"] = (ts.dt.dayofweek >= 5).astype(int)
    result["dry_bulb_c"] = pd.to_numeric(df["dry_bulb_c"], errors="coerce")
    result["relative_humidity_pct"] = pd.to_numeric(df["relative_humidity_pct"], errors="coerce")
    result["ghi_wh_m2"] = pd.to_numeric(df["ghi_wh_m2"], errors="coerce")
    result["cooling_degree"] = (result["dry_bulb_c"] - COMFORT_ASSUMPTION["cooling_target_c"]).clip(lower=0)
    result["heating_degree"] = (20.0 - result["dry_bulb_c"]).clip(lower=0)
    result["occupied_proxy"] = ts.dt.hour.between(OPENING_HOUR, CLOSING_HOUR - 1).astype(int)
    return result.fillna(result.median(numeric_only=True)).fillna(0)


def train_donor_profile_model(
    donor: pd.DataFrame, weather_2024: pd.DataFrame
) -> tuple[HistGradientBoostingRegressor | CalendarShapePredictor, dict[str, Any]]:
    training = donor.merge(weather_2024, on="timestamp", how="inner")
    training = training.dropna(subset=["total_load_kwh", "dry_bulb_c", "relative_humidity_pct", "ghi_wh_m2"])
    training = training.sort_values("timestamp").reset_index(drop=True)
    training["date"] = training["timestamp"].dt.floor("D")
    daily_mean = training.groupby("date")["total_load_kwh"].transform("mean").clip(lower=1e-6)
    training["load_shape_index"] = (training["total_load_kwh"] / daily_mean).clip(lower=0, upper=6)
    features = _calendar_weather_features(training)
    target = training["load_shape_index"]
    split = max(int(len(training) * 0.8), 1)
    model = HistGradientBoostingRegressor(
        max_iter=180,
        learning_rate=0.06,
        max_leaf_nodes=24,
        l2_regularization=0.2,
        random_state=42,
    )
    model.fit(features.iloc[:split], target.iloc[:split])
    prediction = np.clip(model.predict(features.iloc[split:]), 0.02, 6.0)
    actual = training["load_shape_index"].iloc[split:].to_numpy()
    baseline_train = training.iloc[:split].copy()
    baseline_train["hour"] = baseline_train["timestamp"].dt.hour
    baseline_train["is_weekend"] = baseline_train["timestamp"].dt.dayofweek >= 5
    baseline_lookup_series = baseline_train.groupby(["is_weekend", "hour"])["load_shape_index"].mean()
    baseline_lookup = {(bool(key[0]), int(key[1])): float(value) for key, value in baseline_lookup_series.items()}
    validation_calendar = training.iloc[split:][["timestamp"]].copy()
    baseline_prediction = np.array(
        [
            baseline_lookup.get((timestamp.dayofweek >= 5, timestamp.hour), float(target.iloc[:split].mean()))
            for timestamp in validation_calendar["timestamp"]
        ],
        dtype=float,
    )
    model_mae = float(mean_absolute_error(actual, prediction))
    baseline_mae = float(mean_absolute_error(actual, baseline_prediction))
    if model_mae < baseline_mae:
        selected_model: HistGradientBoostingRegressor | CalendarShapePredictor = model
        selected_name = "HistGradientBoostingRegressor normalized donor profile shape model"
        selected_mae = model_mae
        selection_reason = "AI candidate passed the time-holdout improvement gate."
    else:
        selected_model = CalendarShapePredictor(baseline_lookup, float(target.iloc[:split].mean()))
        selected_name = "Validated weekend-hour calendar shape prior"
        selected_mae = baseline_mae
        selection_reason = "AI candidate was rejected because it did not beat the transparent calendar baseline."
    metrics = {
        "selected_model_name": selected_name,
        "selection_reason": selection_reason,
        "training_rows": int(split),
        "validation_rows": int(len(actual)),
        "train_start": str(training["timestamp"].min()),
        "validation_end": str(training["timestamp"].max()),
        "selected_validation_mae_shape_index": round(selected_mae, 4),
        "selected_validation_nmae_pct_of_mean_shape": round(selected_mae / max(actual.mean(), 1e-9) * 100, 2),
        "ai_candidate_name": "HistGradientBoostingRegressor",
        "ai_candidate_validation_mae_shape_index": round(model_mae, 4),
        "ai_candidate_validation_r2": round(float(r2_score(actual, prediction)), 4),
        "calendar_baseline_mae_shape_index": round(baseline_mae, 4),
        "mae_improvement_vs_calendar_baseline_pct": round((baseline_mae - model_mae) / max(baseline_mae, 1e-9) * 100, 2),
        "ai_candidate_selected": bool(model_mae < baseline_mae),
        "target_definition": "hourly main-meter kWh divided by that day's mean kWh",
        "scope": "Validates normalized donor-building temporal shape only; it is not reference case hourly accuracy.",
        "features": list(features.columns),
    }
    return selected_model, metrics


def build_db_hourly_estimate(
    monthly: pd.DataFrame,
    target_weather: pd.DataFrame,
    donor_model: HistGradientBoostingRegressor | CalendarShapePredictor,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    weather = target_weather.copy()
    predicted = np.clip(donor_model.predict(_calendar_weather_features(weather)), 0.02, 6.0)
    weather["donor_shape_prediction"] = predicted
    weather["month"] = weather["timestamp"].dt.strftime("%Y-%m")
    weather["hour"] = weather["timestamp"].dt.hour
    weather["is_weekend"] = weather["timestamp"].dt.dayofweek >= 5
    weather["is_open"] = weather["hour"].between(OPENING_HOUR, CLOSING_HOUR - 1)
    weather["is_occupied"] = weather["is_open"]
    weather["occupancy_fraction_proxy"] = np.where(
        weather["is_open"], np.where(weather["is_weekend"], 0.40, 0.65), 0.03
    )
    cooling = (weather["dry_bulb_c"] - COMFORT_ASSUMPTION["cooling_target_c"]).clip(lower=0)
    heating = (20.0 - weather["dry_bulb_c"]).clip(lower=0)
    weather["thermal_degree"] = cooling + 0.75 * heating
    detail_records: list[pd.DataFrame] = []
    for meter in monthly["meter_name"].unique():
        meter_monthly = monthly.loc[monthly["meter_name"] == meter].set_index("month")
        group = str(meter_monthly["meter_group"].iloc[0])
        for month, month_weather in weather.groupby("month", sort=True):
            if month not in meter_monthly.index:
                continue
            segment = month_weather.copy()
            load_norm = segment["donor_shape_prediction"] / max(segment["donor_shape_prediction"].mean(), 1e-9)
            thermal_norm = segment["thermal_degree"] / max(segment["thermal_degree"].mean(), 1.0)
            if group == "Rooftop HVAC":
                schedule = np.where(segment["is_occupied"], 1.0, 0.28)
                if "MTR-D" in meter.upper():
                    raw_weight = schedule * (0.55 + 0.25 * load_norm + 0.20 * thermal_norm)
                else:
                    raw_weight = schedule * (0.20 + 0.42 * load_norm + 0.38 * thermal_norm)
            else:
                raw_weight = 0.30 + 0.70 * load_norm
            raw_weight = np.asarray(raw_weight, dtype=float)
            raw_weight = np.clip(raw_weight, 1e-6, None)
            monthly_kwh = float(meter_monthly.at[month, "usage_kwh"])
            segment["estimated_kwh"] = monthly_kwh * raw_weight / raw_weight.sum()
            segment["meter_name"] = meter
            segment["meter_group"] = group
            segment["monthly_anchor_kwh"] = monthly_kwh
            segment["estimate_evidence"] = "derived_monthly_constrained_cross_building_transfer"
            detail_records.append(
                segment[
                    [
                        "timestamp",
                        "month",
                        "meter_name",
                        "meter_group",
                        "estimated_kwh",
                        "monthly_anchor_kwh",
                        "dry_bulb_c",
                        "relative_humidity_pct",
                        "ghi_wh_m2",
                        "weather_evidence",
                        "is_open",
                        "is_occupied",
                        "occupancy_fraction_proxy",
                        "estimate_evidence",
                    ]
                ]
            )
    detail = pd.concat(detail_records, ignore_index=True)
    aggregate = (
        detail.pivot_table(index="timestamp", columns="meter_group", values="estimated_kwh", aggfunc="sum", fill_value=0)
        .reset_index()
        .rename(columns={"Rooftop HVAC": "hvac_kwh", "Non-HVAC": "non_hvac_kwh"})
    )
    aggregate = aggregate.merge(
        target_weather[
            ["timestamp", "dry_bulb_c", "relative_humidity_pct", "ghi_wh_m2", "wind_speed_m_s", "weather_evidence"]
        ],
        on="timestamp",
        how="left",
    )
    aggregate["total_kwh"] = aggregate.get("hvac_kwh", 0) + aggregate.get("non_hvac_kwh", 0)
    aggregate["month"] = aggregate["timestamp"].dt.strftime("%Y-%m")
    aggregate["hour"] = aggregate["timestamp"].dt.hour
    aggregate["is_weekend"] = aggregate["timestamp"].dt.dayofweek >= 5
    aggregate["is_open"] = aggregate["hour"].between(OPENING_HOUR, CLOSING_HOUR - 1)
    aggregate["is_occupied"] = aggregate["is_open"]
    aggregate["occupancy_fraction_proxy"] = np.where(
        aggregate["is_open"], np.where(aggregate["is_weekend"], 0.40, 0.65), 0.03
    )
    outdoor = aggregate["dry_bulb_c"].to_numpy(dtype=float)
    open_mask = aggregate["is_open"].to_numpy(dtype=bool)
    comfort_target = np.select(
        [outdoor > 26.0, outdoor < 18.0],
        [COMFORT_ASSUMPTION["cooling_target_c"], COMFORT_ASSUMPTION["heating_target_c"]],
        default=23.0,
    )
    aggregate["indoor_temperature_proxy_c"] = np.where(open_mask, comfort_target, np.clip(outdoor, 15.0, 30.0))
    aggregate["indoor_relative_humidity_proxy_pct"] = np.where(
        open_mask, 50.0, np.clip(aggregate["relative_humidity_pct"].to_numpy(dtype=float), 30.0, 80.0)
    )
    aggregate["indoor_co2_proxy_ppm"] = np.where(
        open_mask, 550.0 + 550.0 * aggregate["occupancy_fraction_proxy"].to_numpy(dtype=float), 450.0
    )
    aggregate["comfort_proxy_pass"] = np.where(
        open_mask,
        aggregate["indoor_temperature_proxy_c"].between(*COMFORT_ASSUMPTION["acceptable_temperature_band_c"])
        & aggregate["indoor_relative_humidity_proxy_pct"].between(*COMFORT_ASSUMPTION["relative_humidity_band_pct"])
        & aggregate["indoor_co2_proxy_ppm"].le(COMFORT_ASSUMPTION["co2_upper_limit_ppm"]),
        True,
    )
    aggregate["comfort_evidence"] = "assumed_comfortable_not_measured"
    aggregate["estimate_evidence"] = "derived_not_measured_hourly"
    reconciliation = (
        detail.groupby(["month", "meter_name"], as_index=False)["estimated_kwh"].sum().merge(
            monthly[["month", "meter_name", "usage_kwh"]], on=["month", "meter_name"], how="left"
        )
    )
    reconciliation["difference_kwh"] = reconciliation["estimated_kwh"] - reconciliation["usage_kwh"]
    reconciliation["status"] = np.where(reconciliation["difference_kwh"].abs() < 1e-6, "PASS", "FAIL")
    return detail, aggregate


def build_scenario_results(hourly: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    base = hourly.copy()
    occupied = base["is_occupied"].astype(bool).to_numpy()
    temperature = base["dry_bulb_c"].to_numpy()
    hvac = base["hvac_kwh"].to_numpy()
    non_hvac = base["non_hvac_kwh"].to_numpy()
    cooling_intensity = np.clip((temperature - 24.0) / 8.0, 0, 1)
    heating_intensity = np.clip((20.0 - temperature) / 10.0, 0, 1)
    thermal_intensity = np.maximum(cooling_intensity, heating_intensity)
    saved_setpoint = hvac * occupied * (0.035 + 0.065 * thermal_intensity)
    saved_schedule = hvac * (~occupied) * 0.35
    saved_led = non_hvac * 0.32 * 0.25
    saved_plug = non_hvac * 0.28 * np.where(occupied, 0.08, 0.25)
    scenario_components = {
        "baseline": np.zeros(len(base)),
        "hvac_setpoint": saved_setpoint,
        "schedule_optimization": saved_schedule,
        "led_retrofit": saved_led,
        "plug_management": saved_plug,
        "combo_package": saved_setpoint + saved_schedule + saved_led + saved_plug,
    }
    definitions = {
        "baseline": ("Baseline", "baseline", 0),
        "hvac_setpoint": ("Comfort-constrained HVAC optimisation", "operational", 8000),
        "schedule_optimization": ("HVAC operating-hours optimisation", "operational", 12000),
        "led_retrofit": ("LED lighting retrofit", "retrofit", 45000),
        "plug_management": ("Plug-load and standby management", "behavioral", 18000),
        "combo_package": ("Combined package (HVAC + LED + operations)", "portfolio", 62000),
    }
    annual_baseline = float(base["total_kwh"].sum())
    detail_rows: list[pd.DataFrame] = []
    summary_rows: list[dict[str, Any]] = []
    for scenario_id, saved_hourly in scenario_components.items():
        name, scenario_type, capex_cny = definitions[scenario_id]
        saved_hourly = np.minimum(np.asarray(saved_hourly), base["total_kwh"].to_numpy() * 0.45)
        annual_saved = float(saved_hourly.sum())
        p10 = annual_saved * 0.75
        p90 = annual_saved * 1.25
        annual_saving_cny = annual_saved * TARIFF["value"]
        payback = capex_cny / annual_saving_cny if capex_cny and annual_saving_cny > 0 else np.nan
        summary_rows.append(
            {
                "scenario_id": scenario_id,
                "scenario_name": name,
                "scenario_type": scenario_type,
                "annual_baseline_kwh": round(annual_baseline, 2),
                "annual_saved_kwh_p10": round(p10, 2),
                "annual_saved_kwh_p50": round(annual_saved, 2),
                "annual_saved_kwh_p90": round(p90, 2),
                "uncertainty_method": "engineering_screening_multipliers_0.75_1.00_1.25",
                "uncertainty_calibration_status": "not_calibrated_as_statistical_quantiles",
                "saving_rate_pct_p50": round(annual_saved / annual_baseline * 100, 2),
                "annual_saving_cny_p10": round(p10 * TARIFF["value"], 2),
                "annual_saving_cny_p50": round(annual_saving_cny, 2),
                "annual_saving_cny_p90": round(p90 * TARIFF["value"], 2),
                "capex_cny_assumption": capex_cny,
                "capex_status": "screening_assumption_needs_supplier_quote" if capex_cny else "not_applicable",
                "simple_payback_years_p50": round(payback, 2) if math.isfinite(payback) else np.nan,
                "avoided_tco2e_maic_p50": round(annual_saved * MAIC_EMISSION_FACTOR["value_kgco2e_per_kwh"] / 1000, 3),
                "carbon_boundary": MAIC_EMISSION_FACTOR["boundary"],
                "result_type": "screening_estimate_with_uncertainty",
            }
        )
        detail = base[["timestamp", "month", "total_kwh"]].copy()
        detail["scenario_id"] = scenario_id
        detail["saved_kwh_p50"] = saved_hourly
        detail["post_scenario_kwh"] = detail["total_kwh"] - saved_hourly
        detail_rows.append(detail)
    return pd.DataFrame(summary_rows), pd.concat(detail_rows, ignore_index=True)


def load_reference_pv(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    raw = pd.read_excel(path, sheet_name="Raw Data")
    raw.columns = ["timestamp", "power_w"]
    raw["timestamp"] = pd.to_datetime(raw["timestamp"], errors="coerce")
    raw["power_w"] = pd.to_numeric(raw["power_w"], errors="coerce").clip(lower=0)
    raw = raw.dropna(subset=["timestamp", "power_w"]).drop_duplicates("timestamp").sort_values("timestamp")
    raw["energy_kwh_5min"] = raw["power_w"] * (5 / 60) / 1000
    hourly = raw.set_index("timestamp")["energy_kwh_5min"].resample("1h").sum().rename("pv_kwh_61_2kwp").reset_index()
    hourly["specific_yield_kwh_per_kwp"] = hourly["pv_kwh_61_2kwp"] / 61.2
    daily = hourly.set_index("timestamp")["pv_kwh_61_2kwp"].resample("1D").sum().reset_index()
    daily["specific_yield_kwh_per_kwp"] = daily["pv_kwh_61_2kwp"] / 61.2
    return hourly, daily


def load_db_pv_monthly(path: Path) -> pd.DataFrame:
    """Read the installed reference case PV monthly generation without altering the source workbook."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    values = {
        "2024-07": ws["B2"].value,
        "2024-08": ws["C2"].value,
        "2024-09": ws["D2"].value,
        "2024-10": ws["E2"].value,
        "2024-11": ws["F2"].value,
        "2024-12": ws["G2"].value,
        "2025-01": ws["B5"].value,
        "2025-02": ws["C5"].value,
        "2025-03": ws["D5"].value,
        "2025-04": ws["E5"].value,
        "2025-05": ws["F5"].value,
        "2025-06": ws["G5"].value,
    }
    wb.close()
    rows = []
    for month, value in values.items():
        rows.append(
            {
                "month": month,
                "pv_generation_kwh": float(value),
                "pv_capacity_kwp": CURRENT_PV["capacity_kwp"],
                "specific_yield_kwh_per_kwp": float(value) / CURRENT_PV["capacity_kwp"],
                "quality_flag": "system_fault_confirmed" if month == "2025-01" else "normal_reported_month",
                "evidence_type": "measured_monthly_pv_generation",
            }
        )
    return pd.DataFrame(rows)


def build_target_pv(reference_hourly: pd.DataFrame, target_index: Iterable[pd.Timestamp], pv_capacity_kwp: float) -> pd.DataFrame:
    ref = reference_hourly.copy()
    ref["month"] = ref["timestamp"].dt.month
    ref["day"] = ref["timestamp"].dt.day
    ref["hour"] = ref["timestamp"].dt.hour
    exact = ref.set_index("timestamp")["specific_yield_kwh_per_kwp"]
    climatology = ref.groupby(["month", "day", "hour"])["specific_yield_kwh_per_kwp"].mean()
    month_hour = ref.groupby(["month", "hour"])["specific_yield_kwh_per_kwp"].mean()
    rows: list[dict[str, Any]] = []
    for timestamp in target_index:
        if timestamp in exact.index:
            specific = float(exact.loc[timestamp])
            evidence = "measured_reference_pv"
        elif (timestamp.month, timestamp.day, timestamp.hour) in climatology.index:
            specific = float(climatology.loc[(timestamp.month, timestamp.day, timestamp.hour)])
            evidence = "reference_pv_climatology"
        else:
            specific = float(month_hour.get((timestamp.month, timestamp.hour), 0.0))
            evidence = "reference_pv_month_hour_fallback"
        rows.append(
            {
                "timestamp": timestamp,
                "specific_yield_kwh_per_kwp": max(specific, 0.0),
                "pv_generation_kwh": max(specific, 0.0) * pv_capacity_kwp,
                "pv_evidence": evidence,
            }
        )
    return pd.DataFrame(rows)


def build_monthly_constrained_db_pv(
    reference_hourly: pd.DataFrame,
    target_index: Iterable[pd.Timestamp],
    monthly_pv: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Transfer only the reference intraday shape and conserve every measured reference case PV month."""
    target = build_target_pv(reference_hourly, target_index, 1.0)
    target["month"] = pd.to_datetime(target["timestamp"]).dt.strftime("%Y-%m")
    anchors = monthly_pv.set_index("month")
    pieces: list[pd.DataFrame] = []
    for month, frame in target.groupby("month", sort=True):
        if month not in anchors.index:
            continue
        segment = frame.copy()
        anchor = float(anchors.at[month, "pv_generation_kwh"])
        weight = segment["pv_generation_kwh"].clip(lower=0).to_numpy(dtype=float)
        if weight.sum() <= 0:
            daylight = segment["timestamp"].dt.hour.between(7, 17).astype(float).to_numpy()
            weight = daylight if daylight.sum() else np.ones(len(segment), dtype=float)
        segment["pv_generation_kwh"] = anchor * weight / weight.sum()
        segment["specific_yield_kwh_per_kwp"] = segment["pv_generation_kwh"] / CURRENT_PV["capacity_kwp"]
        segment["pv_capacity_kwp"] = CURRENT_PV["capacity_kwp"]
        segment["monthly_anchor_kwh"] = anchor
        segment["pv_evidence"] = np.where(
            month == "2025-01",
            "measured_monthly_fault_anchor_plus_reference_intraday_shape",
            "measured_monthly_anchor_plus_reference_intraday_shape",
        )
        pieces.append(segment)
    hourly = pd.concat(pieces, ignore_index=True)
    reconciliation = hourly.groupby("month", as_index=False)["pv_generation_kwh"].sum().merge(
        monthly_pv[["month", "pv_generation_kwh"]].rename(columns={"pv_generation_kwh": "measured_monthly_kwh"}),
        on="month",
        how="left",
    )
    reconciliation["difference_kwh"] = reconciliation["pv_generation_kwh"] - reconciliation["measured_monthly_kwh"]
    reconciliation["status"] = np.where(reconciliation["difference_kwh"].abs() < 1e-6, "PASS", "FAIL")
    return hourly, reconciliation


def estimate_pv_fault_counterfactual(
    reference_hourly: pd.DataFrame,
    target_index: Iterable[pd.Timestamp],
    monthly_pv: pd.DataFrame,
) -> dict[str, Any]:
    """Estimate Jan-2025 no-fault generation while preserving the reported fault value as actual."""
    shape = build_target_pv(reference_hourly, target_index, CURRENT_PV["capacity_kwp"])
    shape["month"] = pd.to_datetime(shape["timestamp"]).dt.strftime("%Y-%m")
    expected = shape.groupby("month")["pv_generation_kwh"].sum()
    actual = monthly_pv.set_index("month")["pv_generation_kwh"]
    calibration_months = [month for month in actual.index if month != "2025-01" and expected.get(month, 0) > 0]
    ratios = [float(actual[month] / expected[month]) for month in calibration_months]
    scale = float(np.median(ratios)) if ratios else 1.0
    no_fault = float(expected.get("2025-01", 0.0) * scale)
    measured = float(actual.get("2025-01", 0.0))
    return {
        "month": "2025-01",
        "measured_fault_generation_kwh": round(measured, 2),
        "estimated_no_fault_generation_kwh": round(no_fault, 2),
        "estimated_generation_loss_kwh": round(max(no_fault - measured, 0.0), 2),
        "method": "reference PV hourly shape scaled by the median performance ratio of the other 11 measured reference case months",
        "result_type": "counterfactual_estimate_not_measured",
    }


def dispatch_no_battery(load: np.ndarray, pv: np.ndarray) -> pd.DataFrame:
    pv_to_load = np.minimum(load, pv)
    grid_to_load = np.maximum(load - pv, 0)
    curtailment = np.maximum(pv - load, 0)
    return pd.DataFrame(
        {
            "grid_to_load_kwh": grid_to_load,
            "grid_to_battery_kwh": np.zeros(len(load)),
            "pv_to_load_kwh": pv_to_load,
            "pv_to_battery_kwh": np.zeros(len(load)),
            "battery_to_load_kwh": np.zeros(len(load)),
            "curtailment_kwh": curtailment,
            "soc_kwh": np.zeros(len(load)),
            "battery_loss_kwh": np.zeros(len(load)),
        }
    )


def dispatch_naive(
    load: np.ndarray,
    pv: np.ndarray,
    hours: np.ndarray,
    capacity_kwh: float,
    power_kw: float,
    eta_charge: float,
    eta_discharge: float,
) -> pd.DataFrame:
    soc = 0.0
    rows: list[dict[str, float]] = []
    horizon = len(load)
    for step, (demand, generation, hour) in enumerate(zip(load, pv, hours, strict=True)):
        pv_to_load = min(demand, generation)
        deficit = max(demand - pv_to_load, 0.0)
        surplus = max(generation - pv_to_load, 0.0)
        pv_charge = min(surplus, power_kw, max((capacity_kwh - soc) / eta_charge, 0.0))
        soc += pv_charge * eta_charge
        surplus -= pv_charge
        grid_charge = 0.0
        # Do not buy grid energy on the final day: all strategies are compared
        # on the same zero-initial/zero-terminal annual accounting boundary.
        if step < horizon - 24 and 1 <= int(hour) <= 5 and soc < 0.8 * capacity_kwh:
            grid_charge = min(power_kw - pv_charge, (0.8 * capacity_kwh - soc) / eta_charge)
            soc += grid_charge * eta_charge
        discharge = 0.0
        if (9 <= int(hour) <= 21 or step >= horizon - 24) and deficit > 0:
            discharge = min(deficit, power_kw, soc * eta_discharge)
            soc -= discharge / eta_discharge
            deficit -= discharge
        loss = (pv_charge + grid_charge) * (1 - eta_charge) + discharge * (1 / eta_discharge - 1)
        rows.append(
            {
                "grid_to_load_kwh": deficit,
                "grid_to_battery_kwh": grid_charge,
                "pv_to_load_kwh": pv_to_load,
                "pv_to_battery_kwh": pv_charge,
                "battery_to_load_kwh": discharge,
                "curtailment_kwh": surplus,
                "soc_kwh": soc,
                "battery_loss_kwh": loss,
            }
        )
    return pd.DataFrame(rows)


def dispatch_loss_aware_lp(
    load: np.ndarray,
    pv: np.ndarray,
    capacity_kwh: float,
    power_kw: float,
    eta_charge: float,
    eta_discharge: float,
    tariff_cny_per_kwh: float,
    degradation_cny_per_kwh_throughput: float = 0.02,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    n = len(load)
    # variable blocks: grid, charge, discharge, curtailment, soc
    idx_g, idx_c, idx_d, idx_u, idx_s = 0, n, 2 * n, 3 * n, 4 * n
    objective = np.zeros(5 * n)
    objective[idx_g : idx_g + n] = tariff_cny_per_kwh
    objective[idx_c : idx_c + n] = degradation_cny_per_kwh_throughput
    objective[idx_d : idx_d + n] = degradation_cny_per_kwh_throughput
    objective[idx_u : idx_u + n] = 1e-6
    surplus = np.maximum(pv - load, 0)
    deficit = np.maximum(load - pv, 0)
    bounds: list[tuple[float, float | None]] = []
    bounds.extend([(0.0, None)] * n)
    bounds.extend([(0.0, float(min(power_kw, value))) for value in surplus])
    bounds.extend([(0.0, float(min(power_kw, value))) for value in deficit])
    bounds.extend([(0.0, float(value)) for value in surplus])
    bounds.extend([(0.0, capacity_kwh)] * n)
    a_eq = lil_matrix((2 * n + 1, 5 * n), dtype=float)
    b_eq = np.zeros(2 * n + 1)
    for t in range(n):
        # grid - charge + discharge - curtailment = load - pv
        a_eq[t, idx_g + t] = 1.0
        a_eq[t, idx_c + t] = -1.0
        a_eq[t, idx_d + t] = 1.0
        a_eq[t, idx_u + t] = -1.0
        b_eq[t] = load[t] - pv[t]
        # soc[t] - soc[t-1] - eta_c*charge + discharge/eta_d = 0
        row = n + t
        a_eq[row, idx_s + t] = 1.0
        if t > 0:
            a_eq[row, idx_s + t - 1] = -1.0
        a_eq[row, idx_c + t] = -eta_charge
        a_eq[row, idx_d + t] = 1.0 / eta_discharge
    # Fair comparison: end SOC equals initial SOC (zero).
    a_eq[2 * n, idx_s + n - 1] = 1.0
    result = linprog(objective, A_eq=a_eq.tocsr(), b_eq=b_eq, bounds=bounds, method="highs")
    if not result.success:
        raise RuntimeError(f"Loss-aware LP failed: {result.message}")
    values = result.x
    grid = values[idx_g : idx_g + n]
    charge = values[idx_c : idx_c + n]
    discharge = values[idx_d : idx_d + n]
    curtailment = values[idx_u : idx_u + n]
    soc = values[idx_s : idx_s + n]
    pv_to_load = np.minimum(load, pv)
    loss = charge * (1 - eta_charge) + discharge * (1 / eta_discharge - 1)
    detail = pd.DataFrame(
        {
            "grid_to_load_kwh": grid,
            "grid_to_battery_kwh": np.zeros(n),
            "pv_to_load_kwh": pv_to_load,
            "pv_to_battery_kwh": charge,
            "battery_to_load_kwh": discharge,
            "curtailment_kwh": curtailment,
            "soc_kwh": soc,
            "battery_loss_kwh": loss,
        }
    )
    meta = {
        "solver": "scipy.optimize.linprog HiGHS",
        "success": bool(result.success),
        "status": int(result.status),
        "message": result.message,
        "objective_cny": float(result.fun),
        "terminal_soc_kwh": float(soc[-1]),
    }
    return detail, meta


def summarize_dispatch(name: str, detail: pd.DataFrame, load: np.ndarray, pv: np.ndarray, capacity_kwh: float) -> dict[str, Any]:
    grid_import = float((detail["grid_to_load_kwh"] + detail["grid_to_battery_kwh"]).sum())
    throughput = float(
        (detail["pv_to_battery_kwh"] + detail["grid_to_battery_kwh"] + detail["battery_to_load_kwh"]).sum()
    )
    return {
        "strategy_id": name,
        "grid_import_kwh": round(grid_import, 2),
        "grid_cost_cny": round(grid_import * TARIFF["value"], 2),
        "pv_generation_kwh": round(float(pv.sum()), 2),
        "battery_loss_kwh": round(float(detail["battery_loss_kwh"].sum()), 2),
        "curtailment_kwh": round(float(detail["curtailment_kwh"].sum()), 2),
        "estimated_export_or_curtailment_kwh": round(float(detail["curtailment_kwh"].sum()), 2),
        "battery_throughput_kwh": round(throughput, 2),
        "equivalent_full_cycles": round(throughput / (2 * capacity_kwh), 2) if capacity_kwh else 0.0,
        "self_sufficiency_pct": round((1 - grid_import / max(float(load.sum()), 1e-9)) * 100, 2),
        "pv_self_consumption_pct": round(
            (1 - float(detail["curtailment_kwh"].sum()) / max(float(pv.sum()), 1e-9)) * 100, 2
        ),
        "terminal_soc_kwh": round(float(detail["soc_kwh"].iloc[-1]), 6),
        "storage_installed": False if name == "no_battery" else "future_hypothetical",
        "result_type": "current_pv_monthly_measured_hourly_estimated"
        if name == "no_battery"
        else "future_storage_screening",
    }


def build_loss_aware_results(
    hourly: pd.DataFrame,
    target_pv: pd.DataFrame,
    assumptions: dict[str, Any],
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    merged = hourly[["timestamp", "total_kwh"]].merge(target_pv, on="timestamp", how="left")
    load = merged["total_kwh"].to_numpy(dtype=float)
    pv = merged["pv_generation_kwh"].fillna(0).to_numpy(dtype=float)
    hours = merged["timestamp"].dt.hour.to_numpy()
    battery = assumptions["future_storage_sandbox"]["battery"]
    no_battery = dispatch_no_battery(load, pv)
    naive = dispatch_naive(
        load,
        pv,
        hours,
        battery["capacity_kwh"],
        battery["power_kw"],
        battery["charge_efficiency"],
        battery["discharge_efficiency"],
    )
    loss_aware, solver_meta = dispatch_loss_aware_lp(
        load,
        pv,
        battery["capacity_kwh"],
        battery["power_kw"],
        battery["charge_efficiency"],
        battery["discharge_efficiency"],
        TARIFF["value"],
        battery["degradation_cny_per_kwh_throughput"],
    )
    details: list[pd.DataFrame] = []
    summaries: list[dict[str, Any]] = []
    for strategy, frame in (("no_battery", no_battery), ("naive_grid_charge", naive), ("loss_aware_lp", loss_aware)):
        frame = frame.copy()
        frame.insert(0, "timestamp", merged["timestamp"].to_numpy())
        frame.insert(1, "strategy_id", strategy)
        frame["load_kwh"] = load
        frame["pv_generation_kwh"] = pv
        frame["grid_import_kwh"] = frame["grid_to_load_kwh"] + frame["grid_to_battery_kwh"]
        details.append(frame)
        summaries.append(summarize_dispatch(strategy, frame, load, pv, battery["capacity_kwh"]))
    return pd.DataFrame(summaries), pd.concat(details, ignore_index=True), solver_meta


def build_pv_storage_sensitivity(hourly: pd.DataFrame, reference_pv: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    target_index = hourly["timestamp"]
    load = hourly["total_kwh"].to_numpy(dtype=float)
    hours = target_index.dt.hour.to_numpy()
    for pv_kwp in (0.0, CURRENT_PV["capacity_kwp"], 150.0):
        target_pv = build_target_pv(reference_pv, target_index, pv_kwp)
        pv = target_pv["pv_generation_kwh"].to_numpy(dtype=float)
        for battery_kwh in (0.0, 150.0, 300.0):
            if battery_kwh == 0:
                detail = dispatch_no_battery(load, pv)
            else:
                detail = dispatch_naive(load, pv, hours, battery_kwh, min(120.0, battery_kwh / 2), 0.95, 0.95)
            summary = summarize_dispatch(f"pv{pv_kwp}_bat{battery_kwh}", detail, load, pv, battery_kwh)
            summary.update({"pv_capacity_kwp": pv_kwp, "battery_capacity_kwh": battery_kwh, "method": "technical_screening"})
            rows.append(summary)
    return pd.DataFrame(rows)


def default_assumptions() -> dict[str, Any]:
    return {
        "model_version": MODEL_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "building": BUILDING,
        "tariff": TARIFF,
        "maic_emission_factor": MAIC_EMISSION_FACTOR,
        "comfort_assumption": COMFORT_ASSUMPTION,
        "hvac_design_basis": HVAC_DESIGN_BASIS,
        "current_energy_system": {
            "pv": CURRENT_PV,
            "storage": {
                "installed": False,
                "status": "confirmed_by_user_2026-08-10",
            },
        },
        "hourly_estimation": {
            "method": "monthly constrained cross-building transfer with measured weather and calendar",
            "weekday_open_hours": "08:00-22:00",
            "weekend_open_hours": "08:00-22:00",
            "opening_hours_status": "user_confirmed_model_assumption_2026-08-10",
            "occupancy_density_status": "estimated_from room use and donor/calendar proxy; not measured",
            "uncertainty_range": "P10/P50/P90 = 0.75/1.00/1.25 of screening savings",
        },
        "scenario_assumptions": {
            "lighting_share_of_non_hvac": 0.32,
            "led_reduction": 0.25,
            "plug_share_of_non_hvac": 0.28,
            "off_hour_hvac_reduction": 0.35,
            "capex_currency": "CNY",
            "capex_status": "screening assumptions; supplier quotes required",
        },
        "future_storage_sandbox": {
            "uses_existing_pv_capacity_kwp": CURRENT_PV["capacity_kwp"],
            "pv_capacity_status": "installed and confirmed; hourly shape is monthly constrained",
            "reference_system_kwp": 61.2,
            "battery": {
                "capacity_kwh": 300.0,
                "power_kw": 120.0,
                "charge_efficiency": 0.95,
                "discharge_efficiency": 0.95,
                "degradation_cny_per_kwh_throughput": 0.02,
                "status": "future hypothetical technical sandbox; no battery is installed",
            },
        },
        "evidence_labels": ["measured", "derived", "assumed", "synthetic_or_sandbox", "missing"],
    }


def build_quality_flags(monthly: pd.DataFrame, hourly_reconciliation: pd.DataFrame) -> pd.DataFrame:
    flags = [
        {
            "flag_id": "DQ-001",
            "severity": "high",
            "scope": "2024-10 MTR-A/MTR-B",
            "issue": "Two non-HVAC cumulative meters show zero increment; monthly building total is only 2,633.67 kWh.",
            "handling": "Preserved as measured data-quality incident; not imputed as normal consumption.",
            "evidence_type": "measured",
        },
        {
            "flag_id": "DQ-002",
            "severity": "medium",
            "scope": "2025-01 to 2025-06 weather",
            "issue": "No matching 2025 campus instrument weather file was provided.",
            "handling": "Uses 2023/2024 month-day-hour climatology and labels every affected hour.",
            "evidence_type": "derived",
        },
        {
            "flag_id": "DQ-003",
            "severity": "high",
            "scope": "reference case hourly load",
            "issue": "No reference case 15-minute/hourly meter readings exist.",
            "handling": "Creates monthly-constrained estimates only; never labels them as measured hourly data.",
            "evidence_type": "missing",
        },
        {
            "flag_id": "DQ-004",
            "severity": "medium",
            "scope": "Donor energy building",
            "issue": "Submeter hierarchy may double-count and some source readings are negative.",
            "handling": "Uses only cleaned main-meter temporal shape; donor absolute kWh never transfers to reference case.",
            "evidence_type": "measured_reference",
        },
        {
            "flag_id": "DQ-005",
            "severity": "medium",
            "scope": "reference case installed PV and metering boundary",
            "issue": "106.14 kWp PV and monthly generation are confirmed, but no inverter-hourly or import/export series is available.",
            "handling": "Preserve measured monthly generation; derive hourly shape and label self-consumption/export as estimates.",
            "evidence_type": "mixed_measured_and_derived",
        },
        {
            "flag_id": "DQ-006",
            "severity": "medium",
            "scope": "Indoor comfort",
            "issue": "No historical temperature, humidity or CO2 records exist.",
            "handling": "Assume comfortable conditions during 08:00-22:00 and generate an explicitly synthetic comfort proxy.",
            "evidence_type": "assumed",
        },
        {
            "flag_id": "DQ-007",
            "severity": "low",
            "scope": "HVAC and envelope design parameters",
            "issue": "2012-2013 drawings describe design intent, not current measured condition; drawing areas conflict with the official area.",
            "handling": "Use equipment/envelope values as priors only; retain 6,231.26 m2 as the canonical current area.",
            "evidence_type": "design_document",
        },
    ]
    failures = hourly_reconciliation.loc[hourly_reconciliation["status"] != "PASS"]
    flags.append(
        {
            "flag_id": "QA-001",
            "severity": "high" if len(failures) else "none",
            "scope": "Hourly monthly conservation",
            "issue": f"{len(failures)} meter-month reconciliation failures.",
            "handling": "PASS: every estimated hourly series sums back to the measured monthly anchor."
            if len(failures) == 0
            else "Stop release and repair allocation.",
            "evidence_type": "validation",
        }
    )
    return pd.DataFrame(flags)


def build_digital_twin_dictionary() -> pd.DataFrame:
    rows = [
        ("BLD_REF", "building", "", "Ningbo Reference Building", "physical_building", "documented", "reference case building profile", "official name; 3 floors; 6231.26 m2; classrooms open 08:00-22:00", "measured occupancy density", "Canonical reference case entity; original source alias: Reference Building A"),
        ("FLR_DB_01", "floor", "BLD_REF", "reference case 1F", "floor", "documented", "reference case building profile", "2047.88 m2", "zone functions and occupancy", "Area documented"),
        ("FLR_DB_02", "floor", "BLD_REF", "reference case 2F", "floor", "documented", "reference case building profile", "2091.69 m2", "zone functions and occupancy", "Area documented"),
        ("FLR_DB_03", "floor", "BLD_REF", "reference case 3F", "floor", "documented", "reference case building profile", "2091.69 m2", "zone functions and occupancy", "Area documented"),
        ("SYS_DB_ELEC", "system", "BLD_REF", "Main metering system", "electrical_system", "measured", "Monthly meter workbook", "MTR-A/MTR-B/MTR-C/MTR-D monthly readings", "circuit topology and hourly logger", "Monthly digital twin anchor"),
        ("MTR_DB_MTR-A", "meter", "SYS_DB_ELEC", "MTR-A", "non_hvac_meter", "measured", "Monthly meter workbook", "monthly kWh; multiplier 40", "end-use mapping", "Source label remains traceable"),
        ("MTR_DB_MTR-B", "meter", "SYS_DB_ELEC", "MTR-B", "non_hvac_meter", "measured", "Monthly meter workbook", "monthly kWh; multiplier 40", "end-use mapping", "Source label remains traceable"),
        ("MTR_DB_MTR-C", "meter", "SYS_DB_ELEC", "MTR-C rooftop HVAC", "hvac_meter", "measured", "Monthly meter workbook", "monthly kWh; multiplier 80", "equipment/COP/control points", "HVAC monthly anchor"),
        ("MTR_DB_MTR-D", "meter", "SYS_DB_ELEC", "MTR-D rooftop HVAC", "hvac_meter", "measured", "Monthly meter workbook", "monthly kWh; multiplier 80", "equipment/COP/control points", "HVAC monthly anchor"),
        ("MODEL_DB_HOURLY", "model", "BLD_REF", "Monthly-constrained hourly estimator", "ai_model", "derived", "reference weather + donor main-meter shape", "estimated hourly profile", "reference case hourly meter validation", "Not measured hourly data"),
        ("SYS_DB_HVAC", "system", "BLD_REF", "VRF HVAC system", "hvac_system", "design_document", "HVAC design drawing", "18 outdoor units; 80 indoor units; 952 kW cooling; rated EER 3.59", "current nameplate and field efficiency", "Design basis calibrated against aggregate HVAC energy"),
        ("SYS_DB_PV", "system", "BLD_REF", "Grid-connected rooftop PV", "pv_system", "measured_and_as_built", "PV generation workbook + as-built drawing", "106.14 kWp; grid connection point 2; monthly generation", "hourly inverter and import/export registers", "Installed current system; January 2025 fault retained"),
        ("SYS_DB_BAT", "system", "BLD_REF", "Battery storage", "storage_system", "confirmed_absent", "Facilities response + user confirmation", "no storage installed", "future design only", "Any battery dispatch is a future sandbox"),
        ("MODEL_DB_COMFORT", "model", "BLD_REF", "Comfort proxy", "indoor_environment_model", "assumed", "User assumption + HVAC design basis", "comfortable during 08:00-22:00", "historical indoor sensors", "Synthetic, never measured"),
        ("MODEL_DB_OPT", "model", "BLD_REF", "Loss-aware LP dispatcher", "optimization_model", "sandbox", "reference case estimated load + installed-PV monthly anchors", "energy balance and terminal SOC", "future battery design and metering boundary", "Battery results are future screening only"),
    ]
    columns = ["entity_id", "entity_level", "parent_id", "entity_name", "entity_type", "evidence_level", "source", "known_fields", "missing_fields", "notes"]
    return pd.DataFrame(rows, columns=columns)


def build_data_request_tracker() -> pd.DataFrame:
    rows = [
        ("P0", "reference case 15-min/hourly electricity", "Unavailable; substitute completed", "Use monthly-constrained 8760 estimate; optional 2-4 week temporary logger for validation", "Supports screening now and later M&V validation"),
        ("P0", "Opening hours", "Assumed by user: every day 08:00-22:00", "Use as the formal operating schedule until timetable/access counts are available", "Anchors occupied/off-hour allocation"),
        ("P0", "Indoor temperature/RH/CO2", "No records; substitute completed", "Use comfortable synthetic proxy; optional low-cost sensors for 2-4 weeks", "Allows comfort-constrained scenarios without claiming measurements"),
        ("P0", "HVAC equipment and COP", "Design schedule available; current condition missing", "Use rated EER 3.59/COP 4.02 with age/part-load uncertainty and aggregate calibration", "Calibrates capacity and control scenarios"),
        ("P0", "PV hourly generation and import/export", "Monthly generation available; hourly boundary missing", "Constrain reference intraday shape to each measured reference case PV month; request inverter/export register if available", "Estimates self-consumption without inventing measurements"),
        ("P1", "2025 Jan-Jun measured weather", "Missing; substitute completed", "Use 2023/2024 month-day-hour campus climatology with hourly evidence labels", "Weather-normalizes the same period as meters"),
        ("P1", "Retrofit supplier quotes in CNY", "Commercially unavailable", "Use P10/P50/P90 screening ranges and replace only when quotes can be disclosed", "Keeps ROI decision boundary explicit"),
        ("P1", "Circuit-to-zone mapping", "Partial from electrical/HVAC drawings", "Use generic HVAC/non-HVAC public groups; replace with an authorized client mapping", "Enables system-level digital twin without fake room precision"),
        ("P1", "Additional roof structural capacity", "Cannot be safely estimated", "Require a structural engineer for any expansion beyond the installed 106.14 kWp", "Prevents unsafe expansion claims"),
        ("P2", "Post-retrofit M&V data", "Future", "Baseline and reporting period under a documented M&V plan", "Verifies savings after implementation"),
    ]
    return pd.DataFrame(rows, columns=["priority", "data_needed", "current_status", "practical_way_to_get_it", "why_it_matters"])


def build_model_readiness_register() -> pd.DataFrame:
    """Describe what each submodel can support and the next evidence upgrade.

    This is deliberately a decision-readiness register, not a single accuracy
    score.  The project contains measured, derived, assumed and sandbox layers,
    so collapsing them into one percentage would overstate field validation.
    """
    rows = [
        (
            "monthly_energy_baseline", "Annual/monthly energy baseline", "Measured + Documented", "Annual analysis",
            "Four monthly meters; October 2024 fault and November includes October", "Annual efficiency, EUI, HVAC share and cost baseline",
            "Cannot reconstruct the true intraday or end-use profile of the anomalous month", "Backend meter logs or November daily records", "P1",
        ),
        (
            "hourly_load_reconstruction", "8,760-hour load reconstruction", "Derived", "Scenario screening",
            "Monthly conservation + campus weather + donor-building intraday shape", "Time-period, PV-coupling and efficiency-scenario screening",
            "Not measured reference case hourly data; cannot support peak/demand commitments", "2–4 weeks of 15-minute reference case main/submeter data", "P0",
        ),
        (
            "comfort_constraint_proxy", "Opening and comfort constraint proxy", "Assumed + Derived", "Constraint testing",
            "Daily 08:00–22:00; 20–26°C, 40–60% RH, CO₂ ≤ 1,000 ppm", "Exclude simulated strategies that clearly compromise comfort",
            "Cannot prove that historical indoor conditions were compliant", "2–4 weeks of temperature/RH/CO₂ in 3–6 representative rooms", "P0",
        ),
        (
            "installed_pv_generation", "Installed PV generation", "Measured monthly", "Annual analysis",
            "106.14 kWp system at grid connection point 2 with 12 measured months", "Annual generation contribution and fault-month identification",
            "Cannot directly determine hourly self-consumption, export or curtailment", "Hourly inverter generation + grid-point import/export", "P0",
        ),
        (
            "pv_self_consumption", "PV self-consumption and grid-import reconstruction", "Derived", "Scenario screening",
            "Measured monthly generation constrains a reference intraday shape paired with estimated load", "Interpret plausible ranges of self-consumption, self-sufficiency and grid imports",
            "The metering boundary and hourly relationship are not site-calibrated", "Synchronous 15-minute inverter and grid-point data", "P0",
        ),
        (
            "retrofit_roi", "Efficiency and ROI scenarios", "Derived + Assumed", "Scenario screening",
            "Engineering rules + 0.75/1.00/1.25 screening multipliers + assumed CAPEX", "Compare measure ranking and sensitivity",
            "P10/P50/P90 are engineering-screening labels, not calibrated probability quantiles; cannot support guaranteed savings", "Equipment verification, supplier quotations and post-implementation M&V", "P0",
        ),
        (
            "future_storage_dispatch", "Future storage dispatch", "Sandbox + Derived", "Technology sandbox",
            "Counterfactual 300 kWh/120 kW battery with energy-balance and terminal-SOC constraints", "Compare rule-based and loss-aware dispatch logic",
            "Uses the Ningbo reference-case tariff of CNY 0.538/kWh; quotations, lifetime data and field-control validation are still missing", "High-frequency load/export data, equipment quotations and degradation parameters", "P0",
        ),
        (
            "local_decision_agent", "Local energy-analysis Agent", "Deterministic + Auditable", "Decision support",
            "Heuristic intent matching + project tools + deterministic calculations", "Queries, combined analysis, scenario recalculation and evidence tracing",
            "Intent-match score is not statistical confidence; the Agent does not connect to or control the BMS", "Expand the regression set with real user questions and evaluate continuously", "P1",
        ),
    ]
    columns = [
        "submodel_id", "submodel_name", "evidence_class", "decision_readiness", "current_basis",
        "appropriate_use", "hard_boundary", "highest_value_upgrade", "upgrade_priority",
    ]
    return pd.DataFrame(rows, columns=columns)


def build_project_summary(
    metadata: dict[str, Any],
    scenario_summary: pd.DataFrame,
    model_metrics: dict[str, Any],
    loss_metrics: pd.DataFrame,
    pv_fault_counterfactual: dict[str, Any] | None = None,
) -> dict[str, Any]:
    annual_total = metadata["annual_total_kwh"]
    annual_hvac = metadata["annual_hvac_kwh"]
    combo = scenario_summary.loc[scenario_summary["scenario_id"] == "combo_package"].iloc[0].to_dict()
    loss_aware = loss_metrics.loc[loss_metrics["strategy_id"] == "loss_aware_lp"].iloc[0].to_dict()
    naive = loss_metrics.loc[loss_metrics["strategy_id"] == "naive_grid_charge"].iloc[0].to_dict()
    current_pv = loss_metrics.loc[loss_metrics["strategy_id"] == "no_battery"].iloc[0].to_dict()
    return {
        "model_version": MODEL_VERSION,
        "project_name": "AI-assisted Auditable Digital Twin Energy Management Platform for Malaysian Buildings and Campuses",
        "real_case": BUILDING,
        "annual_total_kwh": round(annual_total, 2),
        "annual_hvac_kwh": round(annual_hvac, 2),
        "annual_non_hvac_kwh": round(metadata["annual_non_hvac_kwh"], 2),
        "hvac_share_pct": round(annual_hvac / annual_total * 100, 2),
        "annual_eui_kwh_m2": round(annual_total / BUILDING["gross_floor_area_m2"], 2),
        "hvac_eui_kwh_m2": round(annual_hvac / BUILDING["gross_floor_area_m2"], 2),
        "tariff": TARIFF,
        "estimated_annual_bill_cny": round(annual_total * TARIFF["value"], 2),
        "calculated_annual_bill_cny": round(annual_total * TARIFF["value"], 2),
        "opening_hours": {"daily": "08:00-22:00", "status": "user_confirmed_model_assumption_2026-08-10"},
        "comfort_assumption": COMFORT_ASSUMPTION,
        "hvac_design_basis": HVAC_DESIGN_BASIS,
        "current_pv_system": {
            **CURRENT_PV,
            "annual_generation_kwh": current_pv["pv_generation_kwh"],
            "estimated_grid_import_kwh": current_pv["grid_import_kwh"],
            "estimated_self_sufficiency_pct": current_pv["self_sufficiency_pct"],
            "estimated_self_consumption_pct": current_pv["pv_self_consumption_pct"],
            "estimated_export_or_curtailment_kwh": current_pv["estimated_export_or_curtailment_kwh"],
            "metering_boundary": "unknown; self-consumption and grid-import values are screening estimates",
            "january_2025_fault_counterfactual": pv_fault_counterfactual or {},
        },
        "recalibrated_combo_scenario": combo,
        "donor_profile_validation": model_metrics,
        "loss_aware_comparison": {
            "loss_aware_lp": loss_aware,
            "naive_grid_charge": naive,
            "grid_import_reduction_vs_naive_pct": round(
                (naive["grid_import_kwh"] - loss_aware["grid_import_kwh"]) / max(naive["grid_import_kwh"], 1e-9) * 100,
                2,
            ),
            "battery_loss_reduction_vs_naive_pct": round(
                (naive["battery_loss_kwh"] - loss_aware["battery_loss_kwh"]) / max(naive["battery_loss_kwh"], 1e-9) * 100,
                2,
            ),
        },
        "scope_boundary": [
            "Measured: reference case monthly meters and installed-PV monthly generation.",
            "Derived: monthly-constrained load/PV hourly profiles, comfort proxy and cost/EUI calculations.",
            "Assumed: daily 08:00-22:00 opening, comfortable indoor conditions and retrofit CAPEX ranges.",
            "Sandbox: battery dispatch is a future scenario; no storage is installed at reference case.",
        ],
    }


def dataframe_to_markdown(df: pd.DataFrame) -> str:
    """Render a compact Markdown table without the optional tabulate package."""
    columns = [str(column) for column in df.columns]
    lines = ["| " + " | ".join(columns) + " |", "| " + " | ".join(["---"] * len(columns)) + " |"]
    for values in df.itertuples(index=False, name=None):
        cells = []
        for value in values:
            if pd.isna(value):
                text = ""
            else:
                text = str(value).replace("|", "\\|").replace("\n", " ")
            cells.append(text)
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def create_charts(
    monthly: pd.DataFrame,
    monthly_totals: pd.DataFrame,
    hourly: pd.DataFrame,
    scenario_summary: pd.DataFrame,
    loss_metrics: pd.DataFrame,
    loss_detail: pd.DataFrame,
    charts_dir: Path,
) -> list[str]:
    charts_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[str] = []

    fig, ax = plt.subplots(figsize=(11, 5.5))
    ax.plot(monthly_totals["month"], monthly_totals["usage_kwh"], marker="o", linewidth=2.2, color="#0f766e")
    ax.scatter(["2024-10"], [monthly_totals.loc[monthly_totals["month"] == "2024-10", "usage_kwh"].iloc[0]], color="#dc2626", s=80, zorder=3)
    ax.set_title("reference case Monthly Electricity: October 2024 Is a Data-Quality Anomaly")
    ax.set_ylabel("kWh")
    ax.tick_params(axis="x", rotation=45)
    ax.grid(axis="y", alpha=0.25)
    fig.tight_layout()
    path = charts_dir / "monthly-electricity-and-anomaly.png"
    fig.savefig(path, dpi=180)
    plt.close(fig)
    outputs.append(str(path))

    pivot = monthly.pivot_table(index="month", columns="meter_name", values="usage_kwh", aggfunc="sum", fill_value=0)
    fig, ax = plt.subplots(figsize=(11, 6))
    pivot.plot(kind="bar", stacked=True, ax=ax, colormap="viridis")
    ax.set_title("reference case Monthly Electricity by Meter")
    ax.set_xlabel("Month")
    ax.set_ylabel("kWh")
    ax.legend(fontsize=8, loc="upper right")
    fig.tight_layout()
    path = charts_dir / "monthly-meter-breakdown.png"
    fig.savefig(path, dpi=180)
    plt.close(fig)
    outputs.append(str(path))

    plot_df = scenario_summary.loc[scenario_summary["scenario_id"] != "baseline"].copy()
    fig, ax = plt.subplots(figsize=(11, 6))
    x = np.arange(len(plot_df))
    y = plot_df["annual_saved_kwh_p50"].to_numpy()
    yerr = np.vstack([y - plot_df["annual_saved_kwh_p10"].to_numpy(), plot_df["annual_saved_kwh_p90"].to_numpy() - y])
    ax.bar(x, y, color="#0f766e", alpha=0.9)
    ax.errorbar(x, y, yerr=yerr, fmt="none", ecolor="#0f172a", capsize=5)
    ax.set_xticks(x, plot_df["scenario_name"], rotation=18, ha="right")
    ax.set_ylabel("Annual saving (kWh)")
    ax.set_title("Efficiency Scenario P10/P50/P90 Screening Bounds")
    ax.grid(axis="y", alpha=0.2)
    fig.tight_layout()
    path = charts_dir / "scenario-screening-range.png"
    fig.savefig(path, dpi=180)
    plt.close(fig)
    outputs.append(str(path))

    sample = hourly.loc[hourly["timestamp"].between("2024-08-05", "2024-08-11 23:00")]
    fig, ax = plt.subplots(figsize=(12, 5.5))
    ax.plot(sample["timestamp"], sample["total_kwh"], label="Estimated reference case total", color="#0f766e")
    ax.plot(sample["timestamp"], sample["hvac_kwh"], label="Estimated HVAC", color="#f59e0b")
    ax.set_title("Representative Week of Estimated reference case Load (Monthly-Constrained; Not Measured Hourly Data)")
    ax.set_ylabel("kWh/h")
    ax.legend()
    ax.grid(alpha=0.2)
    fig.autofmt_xdate()
    fig.tight_layout()
    path = charts_dir / "representative-hourly-week.png"
    fig.savefig(path, dpi=180)
    plt.close(fig)
    outputs.append(str(path))

    metric_plot = loss_metrics.set_index("strategy_id")[["grid_import_kwh", "battery_loss_kwh"]]
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    metric_plot["grid_import_kwh"].plot(kind="bar", ax=axes[0], color=["#64748b", "#f59e0b", "#0f766e"])
    axes[0].set_title("Annual grid import")
    axes[0].set_ylabel("kWh")
    metric_plot["battery_loss_kwh"].plot(kind="bar", ax=axes[1], color=["#64748b", "#f59e0b", "#0f766e"])
    axes[1].set_title("Battery conversion loss")
    axes[1].set_ylabel("kWh")
    for ax in axes:
        ax.tick_params(axis="x", rotation=18)
        ax.grid(axis="y", alpha=0.2)
    fig.suptitle("Installed 106.14 kWp PV + future 300 kWh battery screening")
    fig.tight_layout()
    path = charts_dir / "annual-dispatch-comparison.png"
    fig.savefig(path, dpi=180)
    plt.close(fig)
    outputs.append(str(path))

    week = loss_detail.loc[
        (loss_detail["strategy_id"] == "loss_aware_lp")
        & loss_detail["timestamp"].between("2024-08-05", "2024-08-11 23:00")
    ]
    fig, ax = plt.subplots(figsize=(12, 5.5))
    ax.plot(week["timestamp"], week["load_kwh"], label="reference case estimated load", color="#0f172a")
    ax.plot(week["timestamp"], week["pv_generation_kwh"], label="PV generation", color="#f59e0b")
    ax.plot(week["timestamp"], week["grid_import_kwh"], label="Grid import", color="#0f766e")
    ax.set_title("Loss-aware LP representative week")
    ax.set_ylabel("kWh/h")
    ax.legend()
    ax.grid(alpha=0.2)
    fig.autofmt_xdate()
    fig.tight_layout()
    path = charts_dir / "representative-week-energy-flow.png"
    fig.savefig(path, dpi=180)
    plt.close(fig)
    outputs.append(str(path))
    return outputs


def build_markdown_reports(
    paths: ProjectPaths,
    summary: dict[str, Any],
    quality_flags: pd.DataFrame,
    scenario_summary: pd.DataFrame,
    data_request: pd.DataFrame,
    model_metrics: dict[str, Any],
    solver_meta: dict[str, Any],
) -> list[str]:
    combo = scenario_summary.loc[scenario_summary["scenario_id"] == "combo_package"].iloc[0]
    diagnosis = f"""# Ningbo Reference Building — Real-Case Data Diagnostic Report

## Official identity and boundary

- Official name: Ningbo Reference Building.
- Original meter-record alias: Reference Building A; retained only for source traceability.
- Gross floor area: 6,231.26 m² across three floors.
- Real-case location: Ningbo, China; target market and carbon scenario: Malaysia.

## Measured monthly baseline

- Annual electricity: {summary['annual_total_kwh']:,.2f} kWh.
- HVAC electricity: {summary['annual_hvac_kwh']:,.2f} kWh, or {summary['hvac_share_pct']:.2f}%.
- Whole-building EUI: {summary['annual_eui_kwh_m2']:.2f} kWh/m²·year.
- HVAC EUI: {summary['hvac_eui_kwh_m2']:.2f} kWh/m²·year.
- Ningbo reference-case billing rule: electricity bill = kWh × CNY 0.538, tax included; calculated annual bill: CNY {summary['calculated_annual_bill_cny']:,.2f}.

## Hourly-estimation method

reference case has no 15-minute or hourly electricity-meter data. Irene uses measured campus weather, the main-meter temporal shape of a donor energy building, the assumed daily 08:00–22:00 opening schedule and calendar features to create an hourly-shape model. Each meter is then scaled month by month so the estimated hourly totals reconcile exactly to measured monthly electricity. The estimates are used only for simulation and are never presented as measured hourly records.

Indoor comfort is represented by a synthetic proxy of 20–26°C, 40–60% RH and CO₂ ≤ 1,000 ppm during opening. The proxy supports comfort-constrained scenario calculations and is not historical indoor measurement.

The normalised donor-building shape passed time-based holdout selection. The selected method is “{model_metrics['selected_model_name']}”, with shape MAE {model_metrics['selected_validation_mae_shape_index']:.3f} and NMAE {model_metrics['selected_validation_nmae_pct_of_mean_shape']:.2f}%. An AI candidate is rejected automatically if it cannot beat the transparent calendar baseline. These metrics do not represent reference case hourly prediction accuracy.

## Critical anomaly

MTR-A and MTR-B cumulative readings did not increase in October 2024, leaving a whole-building monthly total of only 2,633.67 kWh. Facilities confirmed a meter fault and that November includes October consumption. The project retains the original monthly record and labels the boundary in the analysis layer.

## Conclusion

The most reliable current results are the monthly energy baseline, EUI, meter structure, monthly generation of the installed 106.14 kWp PV system and anomaly identification. Hourly profiles, comfort and storage dispatch remain early-stage decision-support outputs with explicit evidence labels.
"""
    scenario_report = f"""# reference case Efficiency, ROI and ESG Screening Report

## Recalibrated combined package

- P50 annual electricity saving: {combo['annual_saved_kwh_p50']:,.2f} kWh.
- P10–P90 screening band: {combo['annual_saved_kwh_p10']:,.2f}–{combo['annual_saved_kwh_p90']:,.2f} kWh.
- P50 saving rate: {combo['saving_rate_pct_p50']:.2f}%.
- P50 annual cost saving: CNY {combo['annual_saving_cny_p50']:,.2f}.
- Ningbo reference-case billing rule: electricity bill = kWh × CNY 0.538, with no time-of-use, demand or other charge components.
- Assumed CAPEX: CNY {combo['capex_cny_assumption']:,.0f}; a supplier quotation is still required.
- P50 simple payback: {combo['simple_payback_years_p50']:.2f} years.
- Avoided emissions under the Malaysia deployment scenario: {combo['avoided_tco2e_maic_p50']:.3f} tCO₂e/year.

## Method improvement

HVAC optimisation is constrained by the comfort range and cannot save energy by sacrificing comfort during opening. Operating-hours optimisation applies only from 22:00 to 08:00, while LED and plug-load measures apply only to their relevant load shares. The combined package aggregates hourly measures with a cap instead of applying one fixed percentage to every hour of the year.

P10/P50/P90 use 0.75/1.00/1.25 engineering-screening multipliers for early-stage sensitivity comparison. They are not statistical quantiles or confidence intervals calibrated from site-error distributions.

## Use limitation

This report is a project-screening study, not a post-implementation M&V result, EPC budget or formal investment decision. It must be recalibrated after CAPEX, occupancy schedules, current HVAC parameters and field-control logic become available.
"""
    loss_report = f"""# reference case PV, Storage and Loss-Aware Optimisation — Technical Note

## Scenario boundary

reference case has an installed 106.14 kWp rooftop PV system at grid connection point 2 and currently has no battery storage. Twelve months of PV generation use measured monthly records; an hourly shape is created by constraining a reference reference PV profile to each month. The January 2025 fault record of 77.5 kWh is retained as measured. The 300 kWh / 120 kW battery exists only in a future scenario sandbox.

## Optimisation method

SciPy HiGHS linear programming minimises electricity-purchase cost and a battery-throughput degradation cost while enforcing hourly energy balance, charge/discharge power, SOC, efficiency and year-end SOC = 0. Solver status: {solver_meta['message']}.

## Result interpretation

The comparison covers installed PV without storage, future rule-based charging and a future loss-aware LP strategy. Grid imports use one consistent definition: grid_to_load + grid_to_battery. Monthly PV generation is measured, while hourly self-consumption, export and grid imports remain estimated. No battery result implies that storage is installed at reference case.
"""
    missing_report = "# reference case Next-Stage Data-Collection Plan\n\n" + dataframe_to_markdown(data_request) + "\n"
    quality_report = "# reference case Data-Quality and Evidence Register\n\n" + dataframe_to_markdown(quality_flags) + "\n"
    files = {
        paths.reports_dir / "building-data-diagnostic.md": diagnosis,
        paths.reports_dir / "efficiency-roi.md": scenario_report,
        paths.reports_dir / "pv-storage-loss-aware.md": loss_report,
        paths.reports_dir / "data-collection-plan.md": missing_report,
        paths.reports_dir / "evidence-register.md": quality_report,
    }
    for path, content in files.items():
        path.write_text(content, encoding="utf-8")
    return [str(path) for path in files]


def build_file_manifest(root: Path) -> pd.DataFrame:
    excluded_directories = {
        ".git",
        ".next",
        ".vercel",
        ".vinext",
        ".wrangler",
        "01_原始资料_只读副本",
        "__pycache__",
        "dist",
        "node_modules",
        "tmp",
    }
    rows: list[dict[str, Any]] = []
    for path in sorted(root.rglob("*")):
        relative_path = path.relative_to(root)
        relative_parts = {part.lower() for part in relative_path.parts}
        is_secret_env = path.name.startswith(".env") and path.name != ".env.example"
        if (
            path.is_file()
            and not relative_parts.intersection(excluded_directories)
            and not is_secret_env
            and path.name != "file_manifest.csv"
        ):
            rows.append(
                {
                    "relative_path": str(relative_path),
                    "extension": path.suffix.lower(),
                    "size_bytes": path.stat().st_size,
                    "sha256": sha256_file(path),
                    "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                }
            )
    return pd.DataFrame(rows)
